import requests
from bs4 import BeautifulSoup
import pandas as pd
from openpyxl import load_workbook

#lists to save the scraped data
titles = []
ratings = []
realease = []
genre = []
duration = []
country = []
description = []

session = requests.Session()
session.max_redirects = 20000

for i in range(307,312):
    print(i)
    page = session.get("https://myflixer.to/top-imdb?type=tv&page=" + str(i))

    soup = BeautifulSoup( page.content , 'html.parser')

    results = soup.find_all("a",{"class":"film-poster-ahref"})

    for result in results:
        
        if result == results[19] and i == 89:
            print('skipped')
            
        elif result == results[12] and i == 193:
            print('skipped')
        
        elif result == results[12] and i == 218:
            print('skipped')
            
        elif result == results[9] and i == 272:
            print('skipped')
        elif result == results[25] and i == 277:
            print('skipped')
            
        elif result == results[9] and i == 289:
            print('skipped')
           
        elif result == results[20] and i == 298:
            print('skipped')
             
        if result == results[3] and i == 307:
            print('skipped')
            
        elif result == results[9] and i == 307:
            print('skipped')
        else:
            anchor = session.get('https://myflixer.to' + result['href'])
            myresult = BeautifulSoup( anchor.content , 'html.parser')
            
            print("--" + str(myresult.find('h2',{'class':'heading-name'}).text))

            #titles
            try:
                titles.append(myresult.find('h2',{'class':'heading-name'}).text)
            except:
                titles.append('none')
                
            #ratings
            try:
                ratings.append(myresult.find('button',{'class':'btn-imdb'}).text)
            except:
                ratings.append('none')
                
            #realease
            try:
                realease.append(myresult.select("div.elements div.row div.col-xl-5.col-lg-6.col-md-8.col-sm-12 div.row-line")[0].text.replace("Released:",'').strip())
            except:
                realease.append('none')
            
            #genre
            try:
                genre1 = myresult.select("div.elements div.row div.col-xl-5.col-lg-6.col-md-8.col-sm-12 div.row-line a")[0].text
                try:
                    genre2 = myresult.select("div.elements div.row div.col-xl-5.col-lg-6.col-md-8.col-sm-12 div.row-line a")[1].text
                except:
                    genre2 = ""
                    
                genre.append(genre1 + ',' + genre2)
            except:
                genre.append('none')
                
            #duration
            try:
                duration.append(myresult.select('div.col-xl-6.col-lg-6.col-md-4.col-sm-12 div.row-line ')[0].text.replace("Duration:",'').replace(' ','').replace('\n', ''))
            except:
                duration.append('none')
            
            #country
            try:
                country.append(myresult.select('div.col-xl-6.col-lg-6.col-md-4.col-sm-12 div.row-line ')[1].text.replace("Country:",'').replace(' ','').replace('\n', ''))
            except:
                country.append('none')
                
            #description
            try:
                description.append(myresult.find("div",{"class":"description"}).text.replace("\n","").strip())
            except:
                description.append('none')


shows = pd.DataFrame({"Title":titles, "Rating":ratings, "Release":realease, "Genre":genre, "Duration":duration, "Country":country, "Description":description}).drop_duplicates()
shows.drop(shows.index[shows['Title'] == 'none'], inplace=True)

filename = 'top_series.xlsx'
book = load_workbook(filename)
writer = pd.ExcelWriter(filename, engine='openpyxl')
writer.book = book
writer.sheets = {ws.title: ws for ws in book.worksheets}

for sheetname in writer.sheets:
    shows.to_excel(writer,sheet_name=sheetname, startrow=writer.sheets[sheetname].max_row, index = False,header= False)

writer.save()
